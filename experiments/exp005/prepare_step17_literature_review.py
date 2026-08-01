import argparse
import math
import os
import re
from copy import copy
from datetime import date
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


DEFAULT_STEP16_DIR = "data/output/starrydata2_step16_result_summary"
DEFAULT_STEP15_DIR = "data/output/starrydata2_step15_pf_zt_error_analysis"
DEFAULT_STEP12_DIR = "data/output/starrydata2_step12_tau_fit"
DEFAULT_STEP13_DIR = "data/output/starrydata2_step13_sigma_validation"
DEFAULT_OUTPUT_DIR = "data/output/starrydata2_step17_literature_review"

STRING_COLUMNS = [
    "sample_key",
    "SID",
    "DOI",
    "doi_url",
    "sample_id",
    "composition",
    "material_system",
]

REQUIRED_TARGET_COLUMNS = [
    "sample_key",
    "step17_review_priority_score",
    "step17_review_priority_tier",
    "step17_review_reason",
]

MANUAL_COLUMNS = [
    "paper_checked_step17",
    "paper_check_date_step17",
    "paper_check_scope_step17",
    "paper_check_note_step17",
    "additive_paper_manual_step17",
    "additive_evidence_paper_step17",
    "additive_confidence_step17",
    "structure_paper_manual_step17",
    "structure_evidence_paper_step17",
    "structure_confidence_step17",
    "np_type_paper_manual_step17",
    "np_basis_paper_manual_step17",
    "np_confidence_paper_step17",
    "sintering_method_paper_manual_step17",
    "sintering_condition_paper_manual_step17",
    "sintering_evidence_paper_step17",
    "sintering_confidence_step17",
    "rare_metal_note_paper_step17",
    "toxicity_note_paper_step17",
    "manual_review_status_step17",
    "manual_review_note_step17",
]


def parse_args():
    parser = argparse.ArgumentParser(description="Prepare Step17 literature review templates and annotation tables.")
    parser.add_argument("--step16_dir", default=DEFAULT_STEP16_DIR)
    parser.add_argument("--step15_dir", default=DEFAULT_STEP15_DIR)
    parser.add_argument("--step12_dir", default=DEFAULT_STEP12_DIR)
    parser.add_argument("--step13_dir", default=DEFAULT_STEP13_DIR)
    parser.add_argument("--output_dir", default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--manual_annotations", default=None)
    parser.add_argument("--top_n_review_targets", type=int, default=300)
    return parser.parse_args()


def dtype_for_existing(path):
    header = pd.read_csv(path, nrows=0)
    return {col: "string" for col in STRING_COLUMNS if col in header.columns}


def read_csv(path, required=False):
    path = Path(path)
    if not path.exists():
        if required:
            raise FileNotFoundError(f"Required input file is missing: {path}")
        return None
    df = pd.read_csv(path, dtype=dtype_for_existing(path), low_memory=False)
    return collapse_duplicate_columns(df)


def collapse_duplicate_columns(df):
    out = df.copy()
    for col in list(out.columns):
        match = re.match(r"^(?P<base>.+)\.\d+$", col)
        if not match:
            continue
        base = match.group("base")
        if base in out.columns:
            out[base] = out[base].where(out[base].map(is_nonempty), out[col])
            out = out.drop(columns=[col])
    return out


def ensure_columns(df, columns, source_name):
    missing = [col for col in columns if col not in df.columns]
    if missing:
        raise ValueError(f"{source_name} is missing required columns: {missing}")


def normalize_text(value):
    if pd.isna(value):
        return ""
    return str(value).strip()


def is_nonempty(value):
    return normalize_text(value) != ""


def is_yes(value):
    return normalize_text(value).lower() in {"yes", "true", "1", "y", "t"}


def prefer_nonempty(*values):
    for value in values:
        if is_nonempty(value):
            return value
    return "unknown"


def source_for(preferred_value, candidates):
    for source, value in candidates:
        if is_nonempty(value) and normalize_text(preferred_value) == normalize_text(value):
            return source
    return "unknown"


def reorder_columns(df, preferred):
    ordered = [col for col in preferred if col in df.columns and col != "doi_url"]
    tail = [col for col in df.columns if col not in ordered and col != "doi_url"]
    if "doi_url" in df.columns:
        tail.append("doi_url")
    return df[ordered + tail]


def make_review_targets(targets, top_n):
    df = targets.copy()
    df = df.sort_values(["step17_review_priority_score", "zt_obs_max_step14"], ascending=[False, False], na_position="last")
    df = df.head(top_n).drop_duplicates(subset=["sample_key"], keep="first")
    preferred = [
        "sample_key",
        "step17_review_priority_score",
        "step17_review_priority_tier",
        "step17_review_reason",
        "DOI",
        "paper_title",
        "sample_id",
        "composition",
        "material_system",
        "n_or_p",
        "n_or_p_basis",
        "n_or_p_step6",
        "n_or_p_basis_step6",
        "n_or_p_confidence_step6",
        "zt_obs_max_step14",
        "zt_pred_max_step14",
        "zt_calc_from_obs_max_step14",
        "zt_pred_vs_obs_mape_step14",
        "zt_pred_vs_calc_mape_step14",
        "classification_case_step15",
        "zt_error_analysis_category_step15",
        "step17_check_additive",
        "step17_check_structure",
        "step17_check_np_type",
        "step17_check_sintering",
        "additive_auto_step9",
        "additive_manual_step9",
        "structure_auto_step9",
        "structure_manual_step9",
        "nanocarbon_keyword_detected_step9",
        "nanocarbon_type_auto_step9",
        "rare_metal_flag_auto_step9",
        "toxicity_flag_auto_step9",
        "sintering_method",
        "sintering_checked",
        "record_checked",
        "doi_url",
    ]
    return reorder_columns(df, preferred)


def make_manual_template(review_targets):
    df = review_targets.copy()
    for col in MANUAL_COLUMNS:
        if col not in df.columns:
            df[col] = ""
    df["paper_checked_step17"] = "no"
    df["paper_check_scope_step17"] = "not checked"
    df["manual_review_status_step17"] = "not_checked"
    preferred = [
        "sample_key",
        "step17_review_priority_tier",
        "step17_review_priority_score",
        "step17_review_reason",
        "paper_checked_step17",
        "paper_check_date_step17",
        "paper_check_scope_step17",
        "paper_check_note_step17",
        "DOI",
        "paper_title",
        "sample_id",
        "composition",
        "material_system",
        "n_or_p",
        "n_or_p_basis",
        "zt_obs_max_step14",
        "zt_pred_max_step14",
        "zt_pred_vs_obs_mape_step14",
        "step17_check_additive",
        "step17_check_structure",
        "step17_check_np_type",
        "step17_check_sintering",
        "additive_paper_manual_step17",
        "additive_evidence_paper_step17",
        "additive_confidence_step17",
        "structure_paper_manual_step17",
        "structure_evidence_paper_step17",
        "structure_confidence_step17",
        "np_type_paper_manual_step17",
        "np_basis_paper_manual_step17",
        "np_confidence_paper_step17",
        "sintering_method_paper_manual_step17",
        "sintering_condition_paper_manual_step17",
        "sintering_evidence_paper_step17",
        "sintering_confidence_step17",
        "rare_metal_note_paper_step17",
        "toxicity_note_paper_step17",
        "manual_review_status_step17",
        "manual_review_note_step17",
        "doi_url",
    ]
    return reorder_columns(df, preferred)


def apply_manual_annotations(template, manual_path):
    applied_samples = 0
    manual_rows = 0
    if not manual_path:
        return template.copy(), applied_samples, manual_rows
    manual_path = Path(manual_path)
    manual = read_csv(manual_path, required=True)
    ensure_columns(manual, ["sample_key"], str(manual_path))
    manual = manual.drop_duplicates(subset=["sample_key"], keep="last")
    manual_rows = len(manual)
    out = template.copy()
    out = out.set_index("sample_key", drop=False)
    manual = manual.set_index("sample_key", drop=False)
    common = [key for key in manual.index if key in out.index]
    for key in common:
        changed = False
        for col in MANUAL_COLUMNS:
            if col in manual.columns:
                value = manual.at[key, col]
                if is_nonempty(value):
                    out.at[key, col] = value
                    changed = True
        if changed:
            applied_samples += 1
    return out.reset_index(drop=True), applied_samples, manual_rows


def add_final_annotation_columns(df):
    out = df.copy()
    for col in MANUAL_COLUMNS:
        if col not in out.columns:
            out[col] = ""

    additive_final = []
    additive_source = []
    additive_conf = []
    structure_final = []
    structure_source = []
    structure_conf = []
    np_final = []
    np_basis = []
    np_source = []
    np_conf = []
    sintering_final = []
    sintering_condition = []
    sintering_source = []
    sintering_conf = []
    sintering_checked_final = []

    for _, row in out.iterrows():
        additive = prefer_nonempty(row.get("additive_paper_manual_step17"), row.get("additive_manual_step9"), row.get("additive_auto_step9"))
        additive_final.append(additive)
        additive_source.append(source_for(additive, [("paper_manual_step17", row.get("additive_paper_manual_step17")), ("manual_step9", row.get("additive_manual_step9")), ("auto_step9", row.get("additive_auto_step9"))]))
        additive_conf.append(row.get("additive_confidence_step17") if is_nonempty(row.get("additive_paper_manual_step17")) else ("unknown" if additive == "unknown" else "auto_or_prior_manual"))

        structure = prefer_nonempty(row.get("structure_paper_manual_step17"), row.get("structure_manual_step9"), row.get("structure_auto_step9"))
        structure_final.append(structure)
        structure_source.append(source_for(structure, [("paper_manual_step17", row.get("structure_paper_manual_step17")), ("manual_step9", row.get("structure_manual_step9")), ("auto_step9", row.get("structure_auto_step9"))]))
        structure_conf.append(row.get("structure_confidence_step17") if is_nonempty(row.get("structure_paper_manual_step17")) else ("unknown" if structure == "unknown" else "auto_or_prior_manual"))

        np_value = prefer_nonempty(row.get("np_type_paper_manual_step17"), row.get("n_or_p"))
        np_final.append(np_value)
        if is_nonempty(row.get("np_type_paper_manual_step17")):
            np_basis.append(prefer_nonempty(row.get("np_basis_paper_manual_step17")))
            np_source.append("paper_manual_step17")
            np_conf.append(prefer_nonempty(row.get("np_confidence_paper_step17")))
        else:
            np_basis.append(prefer_nonempty(row.get("n_or_p_basis")))
            np_source.append("existing_step6_or_step15" if np_value != "unknown" else "unknown")
            np_conf.append(prefer_nonempty(row.get("n_or_p_confidence_step6")))

        check_sintering = is_yes(row.get("step17_check_sintering"))
        sintering_manual = row.get("sintering_method_paper_manual_step17")
        if is_nonempty(sintering_manual):
            sintering = sintering_manual
            sintering_source_value = "paper_manual_step17"
        elif check_sintering:
            sintering = prefer_nonempty(row.get("sintering_method"))
            sintering_source_value = "existing_step7" if sintering != "unknown" else "unknown"
        else:
            sintering = "unknown"
            sintering_source_value = "not_targeted_step17"
        sintering_final.append(sintering)
        sintering_condition.append(row.get("sintering_condition_paper_manual_step17") if is_nonempty(row.get("sintering_condition_paper_manual_step17")) else "")
        sintering_source.append(sintering_source_value)
        sintering_conf.append(row.get("sintering_confidence_step17") if is_nonempty(sintering_manual) else "unknown")
        if is_nonempty(sintering_manual):
            sintering_checked_final.append("yes")
        elif is_yes(row.get("paper_checked_step17")):
            sintering_checked_final.append("not_checked_for_sintering")
        else:
            sintering_checked_final.append("no")

    out["additive_final_step17"] = additive_final
    out["additive_source_step17"] = additive_source
    out["additive_confidence_final_step17"] = additive_conf
    out["structure_final_step17"] = structure_final
    out["structure_source_step17"] = structure_source
    out["structure_confidence_final_step17"] = structure_conf
    out["n_or_p_final_step17"] = np_final
    out["n_or_p_basis_final_step17"] = np_basis
    out["n_or_p_source_step17"] = np_source
    out["n_or_p_confidence_final_step17"] = np_conf
    out["sintering_method_final_step17"] = sintering_final
    out["sintering_condition_final_step17"] = sintering_condition
    out["sintering_source_step17"] = sintering_source
    out["sintering_confidence_final_step17"] = sintering_conf
    out["sintering_checked_final_step17"] = sintering_checked_final
    return out


def make_annotated_samples(annotated):
    preferred = [
        "sample_key",
        "DOI",
        "paper_title",
        "sample_id",
        "composition",
        "material_system",
        "step17_review_priority_score",
        "step17_review_priority_tier",
        "step17_review_reason",
        "paper_checked_step17",
        "paper_check_date_step17",
        "paper_check_scope_step17",
        "manual_review_status_step17",
        "n_or_p",
        "n_or_p_final_step17",
        "n_or_p_basis_final_step17",
        "n_or_p_source_step17",
        "n_or_p_confidence_final_step17",
        "additive_auto_step9",
        "additive_manual_step9",
        "additive_paper_manual_step17",
        "additive_final_step17",
        "additive_source_step17",
        "additive_confidence_final_step17",
        "structure_auto_step9",
        "structure_manual_step9",
        "structure_paper_manual_step17",
        "structure_final_step17",
        "structure_source_step17",
        "structure_confidence_final_step17",
        "sintering_method",
        "sintering_checked",
        "record_checked",
        "sintering_method_paper_manual_step17",
        "sintering_condition_paper_manual_step17",
        "sintering_method_final_step17",
        "sintering_condition_final_step17",
        "sintering_checked_final_step17",
        "sintering_source_step17",
        "sintering_confidence_final_step17",
        "rare_metal_flag_auto_step9",
        "rare_metal_note_paper_step17",
        "toxicity_flag_auto_step9",
        "toxicity_note_paper_step17",
        "zt_obs_max_step14",
        "zt_pred_max_step14",
        "zt_pred_vs_obs_mape_step14",
        "zt_pred_vs_calc_mape_step14",
        "classification_case_step15",
        "zt_error_analysis_category_step15",
        "step17_check_additive",
        "step17_check_structure",
        "step17_check_np_type",
        "step17_check_sintering",
        "manual_review_note_step17",
        "doi_url",
    ]
    return reorder_columns(annotated, preferred)


def make_sintering_targets(annotated):
    mask = annotated["step17_check_sintering"].map(is_yes)
    if "sintering_method_paper_manual_step17" in annotated.columns:
        mask = mask | annotated["sintering_method_paper_manual_step17"].map(is_nonempty)
    preferred = [
        "sample_key",
        "DOI",
        "paper_title",
        "sample_id",
        "composition",
        "material_system",
        "zt_obs_max_step14",
        "zt_pred_max_step14",
        "zt_pred_vs_obs_mape_step14",
        "step17_review_reason",
        "step17_check_sintering",
        "sintering_method_paper_manual_step17",
        "sintering_condition_paper_manual_step17",
        "sintering_evidence_paper_step17",
        "sintering_method_final_step17",
        "sintering_checked_final_step17",
        "manual_review_status_step17",
        "doi_url",
    ]
    return reorder_columns(annotated.loc[mask].copy(), preferred)


def make_tau_eff_ml_base(annotated, tau_fit):
    ann = annotated.copy()
    if tau_fit is not None and "sample_key" in tau_fit.columns:
        tau_cols = [
            "sample_key",
            "tau_eff_step12",
            "log_tau_eff_step12",
            "tau_eff_unit_step12",
            "tau_eff_mode_step12",
            "sigma_fit_log_rmse_step12",
            "sigma_fit_mape_step12",
            "fit_status_step12",
            "fit_note_step12",
        ]
        tau = tau_fit[[col for col in tau_cols if col in tau_fit.columns]].drop_duplicates(subset=["sample_key"], keep="first")
        ann = ann.merge(tau, on="sample_key", how="left", suffixes=("", "_from_step12"))
        for col in tau_cols:
            alt = f"{col}_from_step12"
            if alt in ann.columns:
                if col not in ann.columns:
                    ann[col] = ann[alt]
                else:
                    ann[col] = ann[col].where(ann[col].map(is_nonempty), ann[alt])
                ann = ann.drop(columns=[alt])
    preferred = [
        "sample_key",
        "composition",
        "material_system",
        "n_or_p_final_step17",
        "n_or_p_confidence_final_step17",
        "additive_final_step17",
        "additive_source_step17",
        "additive_confidence_final_step17",
        "structure_final_step17",
        "structure_source_step17",
        "structure_confidence_final_step17",
        "sintering_method_final_step17",
        "sintering_checked_final_step17",
        "nanocarbon_keyword_detected_step9",
        "nanocarbon_type_auto_step9",
        "rare_metal_flag_auto_step9",
        "toxicity_flag_auto_step9",
        "tau_eff_step12",
        "log_tau_eff_step12",
        "tau_eff_unit_step12",
        "tau_eff_mode_step12",
        "sigma_fit_log_rmse_step12",
        "sigma_fit_mape_step12",
        "paper_checked_step17",
        "manual_review_status_step17",
        "doi_url",
    ]
    return reorder_columns(ann, preferred)


def make_status_summary(annotated):
    rows = []
    total = len(annotated)
    for col in [
        "paper_checked_step17",
        "manual_review_status_step17",
        "step17_review_priority_tier",
        "step17_check_additive",
        "step17_check_structure",
        "step17_check_np_type",
        "step17_check_sintering",
        "additive_source_step17",
        "structure_source_step17",
        "n_or_p_source_step17",
        "sintering_source_step17",
        "sintering_checked_final_step17",
    ]:
        if col not in annotated.columns:
            continue
        counts = annotated[col].fillna("not_available").astype(str).replace({"": "not_available"}).value_counts(dropna=False)
        for category, count in counts.items():
            rows.append(
                {
                    "summary_type": col,
                    "category": category,
                    "count": int(count),
                    "fraction": count / total if total else np.nan,
                    "note": "Step17 review status summary",
                }
            )
    return pd.DataFrame(rows)


def count_nonempty(df, col):
    if col not in df.columns:
        return 0
    return int(df[col].map(is_nonempty).sum())


def value_counts_lines(df, col, prefix):
    lines = []
    if col not in df.columns:
        return [f"- {prefix}: column missing"]
    for value, count in df[col].fillna("not_available").astype(str).replace({"": "not_available"}).value_counts().items():
        lines.append(f"- {prefix} {value}: {count}")
    return lines


def make_report(
    input_count,
    review_targets,
    template,
    annotated,
    checked,
    unchecked,
    sintering_targets,
    ml_base,
    manual_path,
    manual_rows,
    manual_applied,
    duplicate_counts,
    n_p_changed_rows,
    sintering_changed_rows,
):
    lines = []
    lines.append("Step17 literature review preparation report")
    lines.append("")
    lines.append(f"input step16_next_step17_review_targets rows: {input_count}")
    lines.append(f"output step17_review_targets rows: {len(review_targets)}")
    lines.append(f"output manual_annotation_template_step17 rows: {len(template)}")
    lines.append(f"output step17_annotated_samples rows: {len(annotated)}")
    lines.append(f"output step17_checked_samples rows: {len(checked)}")
    lines.append(f"output step17_unchecked_samples rows: {len(unchecked)}")
    lines.append(f"output step17_sintering_review_targets rows: {len(sintering_targets)}")
    lines.append(f"output step17_tau_eff_ml_annotation_base rows: {len(ml_base)}")
    lines.append("")
    lines.append("paper_checked_step17 counts:")
    lines.extend(value_counts_lines(annotated, "paper_checked_step17", "paper_checked_step17"))
    lines.append("")
    lines.append("manual_review_status_step17 counts:")
    lines.extend(value_counts_lines(annotated, "manual_review_status_step17", "manual_review_status_step17"))
    lines.append("")
    lines.append("step17_review_priority_tier counts:")
    lines.extend(value_counts_lines(annotated, "step17_review_priority_tier", "step17_review_priority_tier"))
    lines.append("")
    lines.append("check flags:")
    for col in ["step17_check_additive", "step17_check_structure", "step17_check_np_type", "step17_check_sintering"]:
        lines.append(f"- {col}=yes: {int(annotated[col].map(is_yes).sum()) if col in annotated.columns else 0}")
    lines.append("")
    lines.append("final sources:")
    for col in ["additive_source_step17", "structure_source_step17", "n_or_p_source_step17", "sintering_source_step17"]:
        lines.extend(value_counts_lines(annotated, col, col))
    lines.append("")
    lines.append("manual annotations:")
    lines.append(f"- manual_annotations specified: {'yes' if manual_path else 'no'}")
    lines.append(f"- manual_annotations input file: {manual_path or ''}")
    lines.append(f"- manual_annotations rows: {manual_rows}")
    lines.append(f"- manual_annotations applied samples: {manual_applied}")
    lines.append(f"- paper_checked_step17=yes samples: {int(annotated['paper_checked_step17'].map(is_yes).sum()) if 'paper_checked_step17' in annotated.columns else 0}")
    lines.append(f"- additive_paper_manual_step17 filled samples: {count_nonempty(annotated, 'additive_paper_manual_step17')}")
    lines.append(f"- structure_paper_manual_step17 filled samples: {count_nonempty(annotated, 'structure_paper_manual_step17')}")
    lines.append(f"- np_type_paper_manual_step17 filled samples: {count_nonempty(annotated, 'np_type_paper_manual_step17')}")
    lines.append(f"- sintering_method_paper_manual_step17 filled samples: {count_nonempty(annotated, 'sintering_method_paper_manual_step17')}")
    lines.append("")
    lines.append("duplicates:")
    for name, count in duplicate_counts.items():
        lines.append(f"- {name} duplicate sample_key rows: {count}")
    lines.append("")
    lines.append("n/p:")
    lines.append(f"- n_or_p changed rows: {n_p_changed_rows}")
    lines.append("- n_or_p_final_step17 exists: yes")
    lines.append("- n_or_p_source_step17 exists: yes")
    lines.extend(value_counts_lines(annotated, "n_or_p_final_step17", "n_or_p_final_step17"))
    lines.append("")
    lines.append("sintering:")
    lines.append(f"- existing sintering_method/sintering_checked/record_checked changed rows: {sintering_changed_rows}")
    lines.extend(value_counts_lines(annotated, "sintering_method_final_step17", "sintering_method_final_step17"))
    lines.extend(value_counts_lines(annotated, "sintering_checked_final_step17", "sintering_checked_final_step17"))
    lines.append("")
    lines.append("Notes:")
    lines.append("- Step17 did not create new predictions, refit tau_eff, or recalculate PF/ZT.")
    lines.append("- Step17 created an original-paper review template and merged manual annotations when provided.")
    lines.append("- Sintering method review is targeted only to samples requested by Step16, not all samples.")
    return "\n".join(lines) + "\n"


def make_instructions():
    return """# Step17 Manual Review Instructions

## Purpose
Step17 prepares a manual literature-review template for the important samples selected in Step16. It does not infer original-paper content automatically.

## Files to Open
Open `manual_annotation_template_step17.csv` or `manual_annotation_template_step17.xlsx` first. Use `DOI`, `paper_title`, `sample_id`, `composition`, and `doi_url` to locate the original paper.

## What to Check
Check paper evidence for additives, structure information, and n/p type. Record the result in the Step17 paper/manual columns and keep short evidence notes.

## What Not to Check
Do not create new predictions, refit tau_eff, recalculate PF/ZT, or infer paper contents without reading the source. Do not change existing auto-extracted columns directly.

## How to Fill the Template
Set `paper_checked_step17` to `yes` after checking a paper. Use `high`, `medium`, `low`, or `unknown` in confidence columns. Keep automatic Step9 and existing n/p columns unchanged; write paper-confirmed values only in Step17 manual columns.

## How to Re-run the Script After Filling
After editing, save the filled file as `manual_annotation_template_step17_filled.csv`. Re-run:

```bash
python prepare_step17_literature_review.py --manual_annotations data/output/starrydata2_step17_literature_review/manual_annotation_template_step17_filled.csv
```

## Notes for Sintering Check
Check sintering methods only for samples with `step17_check_sintering=yes`. You do not need to investigate sintering methods for all samples.

## Output Files
The main filled-output table is `step17_annotated_samples.csv`. The Step18 input candidate is `step17_tau_eff_ml_annotation_base.csv`. The `doi_url` column is placed near the end of CSV files so commas are less likely to disturb URL handling during manual editing.
"""


def write_excel(path, sheets, manual_sheet_name=None):
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            safe_name = sheet_name[:31]
            df.to_excel(writer, sheet_name=safe_name, index=False)
            ws = writer.book[safe_name]
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                font = copy(cell.font)
                font.bold = True
                cell.font = font
            for column_cells in ws.columns:
                max_len = 0
                col_letter = column_cells[0].column_letter
                header = str(column_cells[0].value)
                for cell in column_cells:
                    text = "" if cell.value is None else str(cell.value)
                    max_len = max(max_len, min(len(text), 60))
                    if header == "doi_url" and cell.row > 1 and text.startswith("http"):
                        cell.hyperlink = text
                        cell.style = "Hyperlink"
                ws.column_dimensions[col_letter].width = max(10, min(max_len + 2, 60))
            if manual_sheet_name and safe_name == manual_sheet_name:
                fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                for cell in ws[1]:
                    if str(cell.value) in MANUAL_COLUMNS:
                        cell.fill = fill
                confidence_cols = [idx + 1 for idx, col in enumerate(df.columns) if "confidence" in col]
                if confidence_cols and len(df) > 0:
                    dv = DataValidation(type="list", formula1='"high,medium,low,unknown"', allow_blank=True)
                    ws.add_data_validation(dv)
                    for col_idx in confidence_cols:
                        letter = ws.cell(row=1, column=col_idx).column_letter
                        dv.add(f"{letter}2:{letter}{len(df)+1}")


def compute_change_counts(review_targets, annotated):
    base = review_targets.set_index("sample_key")
    ann = annotated.set_index("sample_key")
    n_p_changed = 0
    sintering_changed = 0
    for key in ann.index:
        if key not in base.index:
            continue
        for col in ["n_or_p"]:
            if col in base.columns and col in ann.columns and normalize_text(base.at[key, col]) != normalize_text(ann.at[key, col]):
                n_p_changed += 1
                break
        for col in ["sintering_method", "sintering_checked", "record_checked"]:
            if col in base.columns and col in ann.columns and normalize_text(base.at[key, col]) != normalize_text(ann.at[key, col]):
                sintering_changed += 1
                break
    return n_p_changed, sintering_changed


def main():
    args = parse_args()
    step16_dir = Path(args.step16_dir)
    step15_dir = Path(args.step15_dir)
    step12_dir = Path(args.step12_dir)
    step13_dir = Path(args.step13_dir)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    targets_raw = read_csv(step16_dir / "step16_next_step17_review_targets.csv", required=True)
    read_csv(step16_dir / "step16_key_findings_table.csv", required=True)
    if not (step16_dir / "step16_summary_report.txt").exists():
        raise FileNotFoundError(f"Required input file is missing: {step16_dir / 'step16_summary_report.txt'}")
    ensure_columns(targets_raw, REQUIRED_TARGET_COLUMNS, "step16_next_step17_review_targets.csv")

    # Optional inputs are loaded for availability and future compatibility; Step17 outputs can be built from Step16 alone.
    read_csv(step15_dir / "pf_zt_error_samples_step15.csv", required=False)
    read_csv(step15_dir / "manual_review_candidates_step15.csv", required=False)
    read_csv(step15_dir / "sintering_check_candidates_step15.csv", required=False)
    read_csv(step15_dir / "best_candidate_samples_step15.csv", required=False)
    read_csv(step15_dir / "high_zt_missed_and_false_positive_samples_step15.csv", required=False)
    tau_fit = read_csv(step12_dir / "tau_fit_results_step12.csv", required=False)
    read_csv(step13_dir / "tau_validation_primary_results_step13.csv", required=False)

    review_targets = make_review_targets(targets_raw, args.top_n_review_targets)
    template = make_manual_template(review_targets)
    template_applied, manual_applied, manual_rows = apply_manual_annotations(template, args.manual_annotations)
    annotated_full = add_final_annotation_columns(template_applied)
    annotated = make_annotated_samples(annotated_full)
    checked = annotated[annotated["paper_checked_step17"].map(is_yes)].copy()
    unchecked = annotated[~annotated["paper_checked_step17"].map(is_yes)].copy()
    sintering_targets = make_sintering_targets(annotated)
    ml_base = make_tau_eff_ml_base(annotated, tau_fit)
    status_summary = make_status_summary(annotated)

    n_p_changed_rows, sintering_changed_rows = compute_change_counts(review_targets, annotated)
    duplicate_counts = {
        "step17_review_targets.csv": int(review_targets["sample_key"].duplicated().sum()),
        "manual_annotation_template_step17.csv": int(template["sample_key"].duplicated().sum()),
        "step17_annotated_samples.csv": int(annotated["sample_key"].duplicated().sum()),
        "step17_tau_eff_ml_annotation_base.csv": int(ml_base["sample_key"].duplicated().sum()),
    }
    report_text = make_report(
        len(targets_raw),
        review_targets,
        template,
        annotated,
        checked,
        unchecked,
        sintering_targets,
        ml_base,
        args.manual_annotations,
        manual_rows,
        manual_applied,
        duplicate_counts,
        n_p_changed_rows,
        sintering_changed_rows,
    )

    review_targets.to_csv(output_dir / "step17_review_targets.csv", index=False)
    template.to_csv(output_dir / "manual_annotation_template_step17.csv", index=False)
    annotated.to_csv(output_dir / "step17_annotated_samples.csv", index=False)
    status_summary.to_csv(output_dir / "step17_review_status_summary.csv", index=False)
    unchecked.to_csv(output_dir / "step17_unchecked_samples.csv", index=False)
    checked.to_csv(output_dir / "step17_checked_samples.csv", index=False)
    sintering_targets.to_csv(output_dir / "step17_sintering_review_targets.csv", index=False)
    ml_base.to_csv(output_dir / "step17_tau_eff_ml_annotation_base.csv", index=False)
    (output_dir / "step17_manual_review_instructions.md").write_text(make_instructions(), encoding="utf-8")
    (output_dir / "step17_literature_review_report.txt").write_text(report_text, encoding="utf-8")

    write_excel(
        output_dir / "manual_annotation_template_step17.xlsx",
        {"manual_annotation_template": template},
        manual_sheet_name="manual_annotation_template",
    )
    report_df = pd.DataFrame({"review_report": report_text.splitlines()})
    write_excel(
        output_dir / "starrydata2_step17_literature_review.xlsx",
        {
            "review_targets": review_targets,
            "manual_annotation_template": template,
            "annotated_samples": annotated,
            "checked_samples": checked,
            "unchecked_samples": unchecked,
            "sintering_review_targets": sintering_targets,
            "tau_eff_ml_annotation_base": ml_base,
            "review_status_summary": status_summary,
            "review_report": report_df,
        },
        manual_sheet_name="manual_annotation_template",
    )

    sample_key_duplicate_total = sum(duplicate_counts.values())
    print("Done.")
    print("Created:")
    for name in [
        "step17_review_targets.csv",
        "manual_annotation_template_step17.csv",
        "manual_annotation_template_step17.xlsx",
        "step17_annotated_samples.csv",
        "step17_review_status_summary.csv",
        "step17_unchecked_samples.csv",
        "step17_checked_samples.csv",
        "step17_sintering_review_targets.csv",
        "step17_tau_eff_ml_annotation_base.csv",
        "step17_manual_review_instructions.md",
        "step17_literature_review_report.txt",
        "starrydata2_step17_literature_review.xlsx",
    ]:
        print(f"- {name}")
    print("")
    print("Summary:")
    print(f"review targets: {len(review_targets)}")
    print(f"manual template rows: {len(template)}")
    print(f"checked samples: {len(checked)}")
    print(f"unchecked samples: {len(unchecked)}")
    print(f"sintering review targets: {len(sintering_targets)}")
    print(f"manual annotations applied samples: {manual_applied}")
    print(f"additive final paper/manual samples: {int(annotated['additive_source_step17'].isin(['paper_manual_step17','manual_step9']).sum())}")
    print(f"structure final paper/manual samples: {int(annotated['structure_source_step17'].isin(['paper_manual_step17','manual_step9']).sum())}")
    print(f"n/p final paper/manual samples: {int(annotated['n_or_p_source_step17'].eq('paper_manual_step17').sum())}")
    print(f"sintering final paper/manual samples: {int(annotated['sintering_source_step17'].eq('paper_manual_step17').sum())}")
    print(f"tau_eff ML annotation base rows: {len(ml_base)}")
    print(f"sample_key duplicates: {sample_key_duplicate_total}")
    print(f"n/p changed rows: {n_p_changed_rows}")
    print(f"existing sintering columns changed rows: {sintering_changed_rows}")


if __name__ == "__main__":
    main()
